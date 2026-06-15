import csv
from django.contrib import admin
from django.http import HttpResponse
from openpyxl import Workbook
from .models import Course, Notice, GalleryImage, FacultyMember, ContactMessage, CarouselSlide

@admin.action(description="Export selected submissions to CSV")
def export_contact_messages_csv(modeladmin, request, queryset):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="contact_messages.csv"'
    
    writer = csv.writer(response)
    # Write headers
    writer.writerow(['ID', 'Name', 'Email', 'Phone', 'Subject', 'Message', 'Submitted At', 'Is Read'])
    
    # Write data rows
    for msg in queryset:
        # Format date time to local string
        submitted_at_str = msg.submitted_at.strftime('%d-%m-%Y %H:%M:%S') if msg.submitted_at else ""
        writer.writerow([
            msg.id,
            msg.name,
            msg.email,
            msg.phone,
            msg.subject,
            msg.message,
            submitted_at_str,
            "Yes" if msg.is_read else "No"
        ])
        
    return response

@admin.action(description="Export selected submissions to Excel (.xlsx)")
def export_contact_messages_excel(modeladmin, request, queryset):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contact_messages.xlsx"'
    
    # Create openpyxl workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Contact Submissions"
    
    # Write headers
    headers = ['ID', 'Name', 'Email', 'Phone', 'Subject', 'Message', 'Submitted At', 'Is Read']
    ws.append(headers)
    
    # Style header row (optional but makes it look professional)
    from openpyxl.styles import Font, PatternFill
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid') # Academic Navy
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        
    # Write data rows
    for msg in queryset:
        submitted_at_str = msg.submitted_at.strftime('%d-%m-%Y %H:%M:%S') if msg.submitted_at else ""
        ws.append([
            msg.id,
            msg.name,
            msg.email,
            msg.phone,
            msg.subject,
            msg.message,
            submitted_at_str,
            "Yes" if msg.is_read else "No"
        ])
        
    # Auto-adjust column widths for better presentation
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val_str = str(cell.value or '')
            # Cap cell string length to prevent extremely wide columns for message body
            if len(val_str) > 40:
                val_str = val_str[:40] + "..."
            if len(val_str) > max_len:
                max_len = len(val_str)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    wb.save(response)
    return response

@admin.action(description="Mark selected messages as Read")
def mark_as_read(modeladmin, request, queryset):
    queryset.update(is_read=True)

@admin.action(description="Mark selected messages as Unread")
def mark_as_unread(modeladmin, request, queryset):
    queryset.update(is_read=False)

@admin.register(ContactMessage)
class ContactMessageAdmin(admin.ModelAdmin):
    list_display = ('name', 'email', 'phone', 'subject', 'submitted_at', 'is_read')
    list_filter = ('is_read', 'submitted_at')
    search_fields = ('name', 'email', 'subject', 'message')
    actions = [export_contact_messages_csv, export_contact_messages_excel, mark_as_read, mark_as_unread]
    readonly_fields = ('submitted_at',)
    
    # Order by unread messages first, then newest submission
    ordering = ('is_read', '-submitted_at')

@admin.register(Course)
class CourseAdmin(admin.ModelAdmin):
    list_display = ('title', 'code', 'duration_years', 'seats', 'eligibility')
    search_fields = ('title', 'code', 'description')
    list_filter = ('duration_years',)

@admin.register(Notice)
class NoticeAdmin(admin.ModelAdmin):
    list_display = ('title', 'published_date', 'is_active', 'is_urgent')
    list_filter = ('is_active', 'is_urgent', 'published_date')
    search_fields = ('title', 'content')

@admin.register(GalleryImage)
class GalleryImageAdmin(admin.ModelAdmin):
    list_display = ('title', 'category', 'uploaded_at')
    list_filter = ('category', 'uploaded_at')
    search_fields = ('title',)

@admin.register(FacultyMember)
class FacultyMemberAdmin(admin.ModelAdmin):
    list_display = ('name', 'designation', 'department', 'qualification')
    list_filter = ('department',)
    search_fields = ('name', 'designation', 'qualification')

@admin.register(CarouselSlide)
class CarouselSlideAdmin(admin.ModelAdmin):
    list_display = ('title', 'tag', 'order', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('title', 'description')

